import pytest
from openpyxl import Workbook, load_workbook

from app.core.context import AgentContext, set_agent_context


pytestmark = pytest.mark.no_infrastructure


class _FakeArtifactSession:
    """最小 fake async session：使 register_artifact 在无 DB 时也能记录产物。"""

    def __init__(self):
        self._added = []

    async def __aenter__(self):
        return self

    async def __aexit__(self, *exc):
        return False

    def add(self, obj):
        self._added.append(obj)

    async def commit(self):
        # AiArtifact.id 已由 uuid4().hex 在构造时生成，无需数据库回填
        pass


@pytest.fixture
def excel_context(tmp_path, monkeypatch):
    from app.services.ai.tools import document_paths, generated_file_service

    uploads = tmp_path / "uploads"
    uploads.mkdir()
    source = uploads / "sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Month", "Amount"])
    sheet.append(["Jan", 10])
    workbook.save(source)
    monkeypatch.setattr(document_paths, "get_data_base_dir", lambda: str(tmp_path))
    async def workspace_root():
        return str(tmp_path / "agent_workspaces")
    monkeypatch.setattr(document_paths, "resolve_workspace_root", workspace_root)
    monkeypatch.setattr(generated_file_service, "generated_files_root", lambda: tmp_path / "generated")
    # register_artifact 的依赖：workspace 根与 DB 会话均改为测试内无 DB 版本
    async def artifact_workspace_root():
        return tmp_path / "agent_workspaces"
    monkeypatch.setattr(generated_file_service, "_workspace_root", artifact_workspace_root)
    monkeypatch.setattr(
        generated_file_service, "AsyncSessionLocal", lambda: _FakeArtifactSession()
    )
    set_agent_context(AgentContext(
        agent_id="agent", agent_name="Agent", user_id=1, conversation_id="conv",
        authorized_attachment_paths=[str(source)],
    ))
    return source


@pytest.mark.asyncio
async def test_excel_read_range_returns_matrix(excel_context):
    from app.services.ai.tools.excel_document_tool import excel_document_read

    result = await excel_document_read.ainvoke({
        "action": "read_range", "path": str(excel_context),
        "sheet_name": "Sales", "cell_range": "A1:B2",
    })

    assert result["data"]["values"] == [["Month", "Amount"], ["Jan", 10]]


@pytest.mark.asyncio
async def test_excel_write_cells_creates_downloadable_copy(excel_context):
    from app.services.ai.tools.excel_document_tool import excel_document_write

    result = await excel_document_write.ainvoke({
        "action": "write_cells", "path": str(excel_context), "sheet_name": "Sales",
        "cells": [{"address": "B2", "value": 42}], "output_filename": "sales_updated.xlsx",
    })

    assert result["changes"]["written_cells"] == 1
    assert result["artifact"]["download_url"]
    assert load_workbook(excel_context)["Sales"]["B2"].value == 10


@pytest.mark.asyncio
async def test_excel_create_writes_initial_rows(excel_context, tmp_path):
    from app.services.ai.tools.excel_document_tool import excel_document_write

    result = await excel_document_write.ainvoke({
        "action": "create",
        "output_filename": "weather.xlsx",
        "sheet_name": "天气预报",
        "rows": [["时间", "天气"], ["09:00", "晴"]],
    })

    assert result["changes"]["created_workbook"] is True
    assert result["changes"]["appended_rows"] == 2
    output_path = next((tmp_path / "agent_workspaces").rglob("weather.xlsx"))
    workbook = load_workbook(output_path, data_only=False)
    try:
        assert [list(row) for row in workbook["天气预报"].values] == [["时间", "天气"], ["09:00", "晴"]]
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_excel_create_writes_initial_cells(excel_context, tmp_path):
    from app.services.ai.tools.excel_document_tool import excel_document_write

    result = await excel_document_write.ainvoke({
        "action": "create",
        "output_filename": "weather_cells.xlsx",
        "sheet_name": "天气概况",
        "cells": [
            {"address": "A1", "value": "日期"},
            {"address": "B1", "value": "天气"},
            {"address": "A2", "value": "2026-08-27"},
            {"address": "B2", "value": "晴"},
        ],
    })

    assert result["changes"]["created_workbook"] is True
    assert result["changes"]["written_cells"] == 4
    output_path = next((tmp_path / "agent_workspaces").rglob("weather_cells.xlsx"))
    workbook = load_workbook(output_path, data_only=False)
    try:
        assert [list(row) for row in workbook["天气概况"].values] == [
            ["日期", "天气"],
            ["2026-08-27", "晴"],
        ]
    finally:
        workbook.close()
